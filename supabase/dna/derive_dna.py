import openpyxl, collections, csv, sys, statistics
sys.path.insert(0,'/tmp'); import dna_map as M

FILES = {"SIA":("Specific Interest Areas.xlsx","OI"), "WA":("Work Activities.xlsx","IM"),
         "ES":("Essential Skills.xlsx","IM"), "TS":("Transferable Skills.xlsx","IM"),
         "AB":("Abilities.xlsx","IM"), "WS":("Work Styles.xlsx","WI"), "WC":("Work Context.xlsx","CX")}
EXTRA=[("WA","Working with Computers"),("WC","E-Mail"),("WC","Physical Proximity"),
       ("WA","Performing General Physical Activities"),("WC","Spend Time Standing"),
       ("WC","Face-to-Face Discussions with Individuals and Within Teams")]
LAYERS=[("INTEREST",M.INTEREST,3),("ACTIVITY",M.ACTIVITY,4),("SKILL",M.SKILL,5),
        ("ENVIRONMENT",M.ENVIRONMENT,3),("WORKSTYLE",M.WORKSTYLE,2)]
# layer -> file O*NET yang menjadi sumbernya (untuk deteksi layer yang kosong)
LAYER_SRC={"INTEREST":{"SIA"},"ACTIVITY":{"WA"},"SKILL":{"ES","TS","AB","WS"},
           "ENVIRONMENT":{"WC","WA","ES"},"WORKSTYLE":{"WS"}}

need=collections.defaultdict(set)
for _,group,_ in LAYERS:
    for comps in group.values():
        for src,el in comps: need[src].add(el)
for src,el in EXTRA: need[src].add(el)

raw=collections.defaultdict(dict)
for src,(fn,scale) in FILES.items():
    wb=openpyxl.load_workbook(fn, read_only=True); ws=wb.active
    it=ws.iter_rows(values_only=True); hdr=list(next(it)); want=need[src]
    for r in it:
        d=dict(zip(hdr,r))
        if d.get("Scale ID")!=scale: continue
        el=d.get("Element Name")
        if el in want and d.get("Data Value") is not None:
            raw[(src,el)][d["O*NET-SOC Code"]]=float(d["Data Value"])
    wb.close()

norm={}
for key,per in raw.items():
    vs=list(per.values()); lo,hi=min(vs),max(vs); span=(hi-lo) or 1.0
    norm[key]={s:(v-lo)/span*100 for s,v in per.items()}

# okupasi mana yang punya data di tiap file sumber
has_src=collections.defaultdict(set)
for (src,el),per in norm.items():
    for soc in per: has_src[soc].add(src)

related=collections.defaultdict(list)
wb=openpyxl.load_workbook("Related Occupations.xlsx", read_only=True); ws=wb.active
it=ws.iter_rows(values_only=True); hdr=list(next(it))
for r in it:
    d=dict(zip(hdr,r))
    related[d["O*NET-SOC Code"]].append((d["Index"], d["Related O*NET-SOC Code"]))
wb.close()
for k in related: related[k].sort()

all_socs=sorted(has_src)

# Agregasi komponen berbeda per layer, dan ini disengaja:
#
#   MAX  untuk Interest / Activity -- komponennya adalah JALUR
#        ALTERNATIF menuju atribut yang sama. Software developer "membangun"
#        lewat komputer, teknisi lewat mesin; keduanya sah. Kalau dirata-rata,
#        keduanya sama-sama tenggelam karena masing-masing nol di jalur lain.
#
#   MEAN untuk Skill / Work Style / Environment -- komponennya adalah FASET dari satu
#        konstruk. Komunikasi menuntut bicara DAN mendengar DAN menulis, jadi
#        unggul di satu faset saja tidak cukup. Environment juga: satu sinyal
#        yang kebetulan tinggi (semua profesi banyak rapat tatap muka) tidak
#        boleh cukup untuk menyimpulkan lingkungan kerjanya.
AGG_MAX = {"INTEREST","ACTIVITY"}

def raw_scores(soc, group, layer):
    out={}
    for code,comps in group.items():
        vs=[norm[c][soc] for c in comps if soc in norm.get(c,{})]
        if vs: out[code] = max(vs) if layer in AGG_MAX else statistics.fmean(vs)
    return out

def env_rules(soc, base):
    g=lambda k: norm.get(k,{}).get(soc)
    inv=lambda v: (100-v) if v is not None else None
    parts=[g(("WA","Working with Computers")), g(("WC","E-Mail")),
           inv(g(("WC","Physical Proximity"))), inv(g(("WA","Performing General Physical Activities"))),
           inv(g(("WC","Spend Time Standing")))]
    parts=[p for p in parts if p is not None]
    if not parts: return
    remote=statistics.fmean(parts); base["ENV_REMOTE"]=remote
    hp=[x for x in (remote, base.get("ENV_KANTOR"), g(("WC","Face-to-Face Discussions with Individuals and Within Teams"))) if x is not None]
    base["ENV_HYBRID"]=statistics.fmean(hp)

# ---- tahap 1: skor mentah, per layer, hanya untuk okupasi yang datanya ada
per_soc={}
for soc in all_socs:
    layers={}
    for lname,group,_ in LAYERS:
        if not (LAYER_SRC[lname] & has_src[soc]): continue
        d=raw_scores(soc, group, lname)
        if lname=="ENVIRONMENT": env_rules(soc, d)
        if d: layers[lname]=d
    per_soc[soc]=layers

# ---- tahap 2: warisi layer yang kosong dari profesi terkait terdekat yang lengkap
inherit={}
for soc in all_socs:
    for lname,_,_ in LAYERS:
        if lname in per_soc[soc]: continue
        for _,rel in related.get(soc,[]):
            if rel in per_soc and lname in per_soc[rel]:
                per_soc[soc][lname]=dict(per_soc[rel][lname])
                inherit.setdefault(soc,{})[lname]=rel
                break

complete=[s for s in all_socs if all(l in per_soc[s] for l,_,_ in LAYERS)]
print(f"  okupasi lengkap: {len(complete)} / {len(all_socs)} | mewarisi minimal 1 layer: {len(inherit)}", file=sys.stderr)

# ---- tahap 3: normalisasi TIAP ATRIBUT lintas okupasi (0..100)
#      Tanpa ini, atribut dengan sedikit komponen (mis. INT_PENDIDIKAN = 1 area)
#      selalu unggul atas atribut dengan banyak komponen yang saling menetralkan.
by_attr=collections.defaultdict(dict)
for soc in complete:
    for lname,group,_ in LAYERS:
        for code,v in per_soc[soc][lname].items(): by_attr[code][soc]=v
scaled=collections.defaultdict(dict)
for code,per in by_attr.items():
    vs=sorted(per.values()); lo,hi=vs[0],vs[-1]; span=(hi-lo) or 1.0
    for soc,v in per.items(): scaled[soc][code]=(v-lo)/span*100

rows=[]
for soc in complete:
    for lname,group,topn in LAYERS:
        codes=[c for c in per_soc[soc][lname]]
        order=sorted(codes, key=lambda c:-scaled[soc][c])
        for i,code in enumerate(order,1):
            src = "inherited:"+inherit.get(soc,{}).get(lname) if inherit.get(soc,{}).get(lname) else "onet"
            if code in ("ENV_REMOTE","ENV_HYBRID") and src=="onet": src="rule"
            rows.append((soc,code,lname,round(scaled[soc][code],2),i,i<=topn,src))

with open("/tmp/career_dna.csv","w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(["soc_code","attribute_code","layer","score","rank_in_layer","is_dominant","dna_source"])
    for r in rows: w.writerow(r)
print(f"  baris: {len(rows)} | okupasi: {len(complete)}", file=sys.stderr)
